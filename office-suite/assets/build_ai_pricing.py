"""
Build AI Model API Pricing spreadsheet.
Data source: costgoat.com / awesomeagents.ai, verified March 14, 2026.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter

OUTPUT = "ai_api_pricing.xlsx"

# ---------------------------------------------------------------------------
# Data: (provider, model, tier, input_usd, output_usd, context_k, cache_note, notes)
# Prices per 1M tokens in USD
# ---------------------------------------------------------------------------
DATA = [
    # OpenAI
    ("OpenAI", "GPT-5.4",         "Flagship",  2.50,  15.00, "272K", "~$0.25 cached",  "OpenAI flagship; 2x price >272K ctx"),
    ("OpenAI", "GPT-5.4 Pro",     "Ultra",    30.00, 180.00, "272K", "-",               "Ultra-premium; researchers/agents"),
    ("OpenAI", "GPT-5.2",         "Advanced",  1.75,  14.00, "400K", "~$0.18 cached",  "Strong reasoning"),
    ("OpenAI", "GPT-5.2 Pro",     "Ultra",    21.00, 168.00, "400K", "-",               "Pro variant"),
    ("OpenAI", "GPT-5.1",         "Advanced",  1.25,  10.00, "400K", "~$0.13 cached",  ""),
    ("OpenAI", "GPT-5",           "Advanced",  1.25,  10.00, "400K", "~$0.13 cached",  ""),
    ("OpenAI", "GPT-5 Pro",       "Ultra",    15.00, 120.00, "400K", "-",               ""),
    ("OpenAI", "GPT-5 Mini",      "Budget",    0.25,   2.00, "400K", "~$0.025 cached", ""),
    ("OpenAI", "GPT-4.1",         "Mid",       2.00,   8.00,   "1M", "~$0.20 cached",  "Long context, strong code"),
    ("OpenAI", "GPT-4.1 mini",    "Budget",    0.40,   1.60, "128K", "~$0.04 cached",  ""),
    ("OpenAI", "GPT-4o mini",     "Budget",    0.15,   0.60, "128K", "~$0.015 cached", "Legacy, still popular"),
    ("OpenAI", "o4-mini",         "Reasoning", 1.10,   4.40, "128K", "~$0.11 cached",  "Affordable reasoning model"),
    ("OpenAI", "o3",              "Reasoning", 2.00,   8.00, "200K", "-",               "Heavy reasoning"),
    ("OpenAI", "o3-pro",          "Reasoning",20.00,  80.00, "200K", "-",               ""),
    ("OpenAI", "o1",              "Reasoning",15.00,  60.00, "128K", "-",               "Legacy reasoning"),
    # Anthropic
    ("Anthropic", "Claude Opus 4.6",   "Flagship", 5.00, 25.00, "200K", "~$0.50 cached", "Top reasoning & agentic; 2x >200K"),
    ("Anthropic", "Claude Opus 4.5",   "Flagship", 5.00, 25.00, "200K", "~$0.50 cached", ""),
    ("Anthropic", "Claude Sonnet 4.6", "Mid",      3.00, 15.00,   "1M", "~$0.30 cached", "Best mid-tier for code"),
    ("Anthropic", "Claude Sonnet 4.5", "Mid",      3.00, 15.00,   "1M", "~$0.30 cached", ""),
    ("Anthropic", "Claude Haiku 4.5",  "Budget",   0.80,  4.00, "200K", "~$0.08 cached", "Fast & cheap"),
    # Google
    ("Google", "Gemini 3.1 Pro",       "Flagship", 2.00, 12.00,   "1M", "~$0.20 cached", "Google flagship; 2x >200K"),
    ("Google", "Gemini 3 Pro",         "Flagship", 2.00, 12.00,   "1M", "~$0.20 cached", ""),
    ("Google", "Gemini 3 Flash",       "Mid",      0.50,  3.00,   "1M", "~$0.05 cached", "Latest Flash"),
    ("Google", "Gemini 2.5 Pro",       "Mid",      1.25, 10.00,   "1M", "~$0.13 cached", "Long context"),
    ("Google", "Gemini 2.5 Flash",     "Budget",   0.30,  2.50,   "1M", "~$0.03 cached", "Solid mid-tier"),
    ("Google", "Gemini 2.5 Flash-Lite","Budget",   0.10,  0.40,   "1M", "-",              "Cheapest Google option"),
    # DeepSeek
    ("DeepSeek", "DeepSeek V3.2",      "Flagship", 0.28,  0.42, "164K", "$0.028 auto",   "Best value; auto cache 90% off"),
    ("DeepSeek", "DeepSeek R1",        "Reasoning",0.55,  2.19, "164K", "$0.055 auto",   "Reasoning; open-weight"),
    # Mistral
    ("Mistral", "Mistral Large 3",     "Flagship", 0.50,  1.50, "128K", "-",              "Strong multilingual"),
    ("Mistral", "Mistral Small 3.1",   "Mid",      0.10,  0.30, "128K", "-",              "Structured tasks"),
    ("Mistral", "Mistral Nemo",        "Budget",   0.02,  0.04, "128K", "-",              "Cheapest commercial API"),
    ("Mistral", "Ministral 3B",        "Budget",   0.04,  0.04, "128K", "-",              "Edge/mobile"),
    # xAI
    ("xAI",     "Grok 4.20",           "Flagship", 2.00,  6.00, "256K", "-",              "Multi-agent capable"),
    ("xAI",     "Grok 4.1 Fast",       "Budget",   0.20,  0.50,   "2M", "$0.05 cached",  "Largest context window"),
    # MoonshotAI
    ("MoonshotAI", "Kimi K2.5",        "Flagship", 0.45,  2.20, "262K", "-",              "Strong value score"),
    ("MoonshotAI", "Kimi K2 Thinking", "Reasoning",0.47,  2.00, "131K", "-",              ""),
    # MiniMax
    ("MiniMax",  "MiniMax M2.5",       "Flagship", 0.27,  0.95, "197K", "-",              ""),
    ("MiniMax",  "MiniMax M2.1",       "Mid",      0.27,  0.95, "197K", "-",              ""),
    # Zhipu AI
    ("Zhipu AI", "GLM-5",              "Flagship", 0.72,  2.30, "203K", "-",              ""),
    ("Zhipu AI", "GLM-4.7",            "Mid",      0.38,  1.98, "203K", "-",              ""),
    # Groq (inference provider, hosts open-weight models)
    ("Groq",    "Llama 4 Maverick",    "Mid",      0.20,  0.60, "128K", "-",              "MoE 17Bx128E; 840 tok/s"),
    ("Groq",    "Llama 4 Scout",       "Budget",   0.11,  0.34, "128K", "-",              "MoE 17Bx16E"),
    ("Groq",    "Llama 3.3 70B",       "Mid",      0.59,  0.79, "128K", "-",              "Open-weight workhorse"),
    ("Groq",    "Llama 3.1 8B",        "Budget",   0.05,  0.08, "128K", "-",              ""),
    ("Groq",    "Qwen3 32B",           "Mid",      0.29,  0.59, "131K", "-",              "Strong multilingual"),
    ("Groq",    "Kimi K2",             "Mid",      1.00,  3.00, "256K", "-",              "via Groq"),
]

# Provider → fill color (hex, no #)
PROVIDER_COLORS = {
    "OpenAI":     "DDEEFF",
    "Anthropic":  "FFE8D6",
    "Google":     "E2F0D9",
    "DeepSeek":   "EDE7F6",
    "Mistral":    "FFF9C4",
    "xAI":        "F3F3F3",
    "MoonshotAI": "E0F4FF",
    "MiniMax":    "FCE4EC",
    "Zhipu AI":   "E8F5E9",
    "Groq":       "FFF3E0",
}

TIER_COLORS = {
    "Flagship":  "FF9800",
    "Ultra":     "E91E63",
    "Advanced":  "3F51B5",
    "Mid":       "009688",
    "Budget":    "4CAF50",
    "Reasoning": "9C27B0",
}

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def header_border():
    s = Side(style='medium', color='888888')
    return Border(left=s, right=s, top=s, bottom=s)

def build():
    wb = Workbook()

    # ── Sheet 1: Full Pricing Table ──────────────────────────────────────────
    ws = wb.active
    ws.title = "全览 All Models"

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "AI 模型 API 定价总览  ·  AI Model API Pricing (March 2026)"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="1A1A2E")
    title_cell.fill = PatternFill("solid", fgColor="1A1A2E")
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    src = ws["A2"]
    src.value = "数据来源 Sources: costgoat.com · awesomeagents.ai  |  价格单位: USD / 百万 Token (1M tokens)  |  每50%批量折扣适用于所有主流服务商"
    src.font = Font(name="Arial", size=9, italic=True, color="666666")
    src.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Column headers ───────────────────────────────────────────────────────
    HDR_ROW = 3
    headers = [
        ("A", "服务商\nProvider",     16),
        ("B", "模型\nModel",          22),
        ("C", "档次\nTier",           10),
        ("D", "输入价格\nInput ($/1M)",12),
        ("E", "输出价格\nOutput ($/1M)",13),
        ("F", "上下文\nContext",       10),
        ("G", "缓存价格\nCached Input",14),
        ("H", "批量价格\nBatch Input",13),
        ("I", "备注\nNotes",          30),
    ]
    hdr_fill   = PatternFill("solid", fgColor="1A1A2E")
    hdr_font   = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_letter, label, width in headers:
        cell = ws[f"{col_letter}{HDR_ROW}"]
        cell.value = label
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = header_border()
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[HDR_ROW].height = 30

    # ── Data rows ────────────────────────────────────────────────────────────
    row = HDR_ROW + 1
    prev_provider = None
    price_rows = []  # track row numbers for conditional formatting

    for provider, model, tier, inp, out, ctx, cache, notes in DATA:
        bg = PROVIDER_COLORS.get(provider, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=bg)
        tier_color = TIER_COLORS.get(tier, "999999")

        # Provider cell (show only when provider changes)
        prov_cell = ws[f"A{row}"]
        if provider != prev_provider:
            prov_cell.value = provider
            prov_cell.font = Font(name="Arial", size=9, bold=True, color="1A1A2E")
        else:
            prov_cell.value = ""
        prov_cell.fill = row_fill
        prov_cell.alignment = Alignment(horizontal="center", vertical="center")
        prov_cell.border = thin_border()

        # Model
        model_cell = ws[f"B{row}"]
        model_cell.value = model
        model_cell.font = Font(name="Arial", size=9, color="1A1A2E")
        model_cell.fill = row_fill
        model_cell.alignment = Alignment(vertical="center")
        model_cell.border = thin_border()

        # Tier badge (colored text)
        tier_cell = ws[f"C{row}"]
        tier_cell.value = tier
        tier_cell.font = Font(name="Arial", size=8, bold=True, color=tier_color)
        tier_cell.fill = row_fill
        tier_cell.alignment = Alignment(horizontal="center", vertical="center")
        tier_cell.border = thin_border()

        # Input price
        inp_cell = ws[f"D{row}"]
        inp_cell.value = inp
        inp_cell.number_format = '"$"#,##0.00'
        inp_cell.font = Font(name="Arial", size=9, color="0000CC")  # blue = hardcoded input
        inp_cell.fill = row_fill
        inp_cell.alignment = Alignment(horizontal="right", vertical="center")
        inp_cell.border = thin_border()

        # Output price
        out_cell = ws[f"E{row}"]
        out_cell.value = out
        out_cell.number_format = '"$"#,##0.00'
        out_cell.font = Font(name="Arial", size=9, color="0000CC")
        out_cell.fill = row_fill
        out_cell.alignment = Alignment(horizontal="right", vertical="center")
        out_cell.border = thin_border()

        # Context
        ctx_cell = ws[f"F{row}"]
        ctx_cell.value = ctx
        ctx_cell.font = Font(name="Arial", size=9, color="444444")
        ctx_cell.fill = row_fill
        ctx_cell.alignment = Alignment(horizontal="center", vertical="center")
        ctx_cell.border = thin_border()

        # Cache
        cache_cell = ws[f"G{row}"]
        cache_cell.value = cache
        cache_cell.font = Font(name="Arial", size=8, color="555555")
        cache_cell.fill = row_fill
        cache_cell.alignment = Alignment(horizontal="center", vertical="center")
        cache_cell.border = thin_border()

        # Batch price (formula: 50% of input)
        batch_cell = ws[f"H{row}"]
        batch_cell.value = f"=D{row}*0.5"
        batch_cell.number_format = '"$"#,##0.00'
        batch_cell.font = Font(name="Arial", size=9, color="008000")  # green = formula
        batch_cell.fill = row_fill
        batch_cell.alignment = Alignment(horizontal="right", vertical="center")
        batch_cell.border = thin_border()

        # Notes
        notes_cell = ws[f"I{row}"]
        notes_cell.value = notes
        notes_cell.font = Font(name="Arial", size=8, color="555555", italic=True)
        notes_cell.fill = row_fill
        notes_cell.alignment = Alignment(vertical="center", wrap_text=True)
        notes_cell.border = thin_border()

        ws.row_dimensions[row].height = 16
        price_rows.append(row)
        prev_provider = provider
        row += 1

    # ── Conditional formatting: color scale on Input (D) and Output (E) ──────
    if price_rows:
        d_range = f"D{price_rows[0]}:D{price_rows[-1]}"
        e_range = f"E{price_rows[0]}:E{price_rows[-1]}"
        scale = ColorScaleRule(
            start_type="min", start_color="63BE7B",   # green = cheap
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",       # red = expensive
        )
        ws.conditional_formatting.add(d_range, scale)
        ws.conditional_formatting.add(e_range, scale)

    # ── Freeze panes & filter ────────────────────────────────────────────────
    ws.freeze_panes = f"A{HDR_ROW + 1}"
    ws.auto_filter.ref = f"A{HDR_ROW}:I{row - 1}"


    # ── Sheet 2: Summary / Quick Comparison ─────────────────────────────────
    ws2 = wb.create_sheet("快速对比 Summary")

    ws2.merge_cells("A1:G1")
    t = ws2["A1"]
    t.value = "AI API 价格快速对比 · Quick Comparison Summary"
    t.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1A1A2E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    # By-provider stats table
    ws2["A3"].value = "服务商 / Provider"
    ws2["B3"].value = "最低输入 Min Input"
    ws2["C3"].value = "最低输出 Min Output"
    ws2["D3"].value = "最高输入 Max Input"
    ws2["E3"].value = "最高输出 Max Output"
    ws2["F3"].value = "旗舰型号 Flagship Model"
    ws2["G3"].value = "旗舰输出价 Flagship Out"

    for col in "ABCDEFG":
        c = ws2[f"{col}3"]
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="37474F")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = header_border()
    ws2.row_dimensions[3].height = 28

    # Group data by provider for summary
    from collections import defaultdict
    provider_data = defaultdict(list)
    flagship_map = {}
    for provider, model, tier, inp, out, ctx, cache, notes in DATA:
        provider_data[provider].append((inp, out, model, tier))
        if tier in ("Flagship", "Ultra") and provider not in flagship_map:
            flagship_map[provider] = (model, out)

    sum_row = 4
    for provider in PROVIDER_COLORS.keys():
        if provider not in provider_data:
            continue
        rows = provider_data[provider]
        inputs = [r[0] for r in rows]
        outputs = [r[1] for r in rows]
        fg_model, fg_out = flagship_map.get(provider, ("—", "—"))
        bg = PROVIDER_COLORS[provider]
        rfill = PatternFill("solid", fgColor=bg)

        def sc(col, val, fmt=None, is_price=False):
            c = ws2[f"{col}{sum_row}"]
            c.value = val
            c.fill = rfill
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Arial", size=9,
                          color="0000CC" if is_price else "1A1A2E")
            if fmt:
                c.number_format = fmt

        sc("A", provider)
        sc("B", min(inputs),  '"$"#,##0.00', True)
        sc("C", min(outputs), '"$"#,##0.00', True)
        sc("D", max(inputs),  '"$"#,##0.00', True)
        sc("E", max(outputs), '"$"#,##0.00', True)
        sc("F", fg_model)
        sc("G", fg_out if isinstance(fg_out, str) else fg_out, '"$"#,##0.00', not isinstance(fg_out, str))
        ws2.row_dimensions[sum_row].height = 16
        sum_row += 1

    # Column widths
    for col, w in [("A",14),("B",14),("C",15),("D",14),("E",15),("F",22),("G",16)]:
        ws2.column_dimensions[col].width = w

    # ── Cost Calculator section ──────────────────────────────────────────────
    calc_start = sum_row + 2
    ws2.merge_cells(f"A{calc_start}:G{calc_start}")
    ch = ws2[f"A{calc_start}"]
    ch.value = "💰 成本估算器 Cost Estimator  (修改黄色单元格 · Edit yellow cells)"
    ch.font = Font(name="Arial", size=10, bold=True, color="1A1A2E")
    ch.fill = PatternFill("solid", fgColor="FFF9C4")
    ch.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[calc_start].height = 20

    labels = [
        ("每日输入 Token 数 (百万) · Daily Input Tokens (M)", 1.0, "B"),
        ("每日输出 Token 数 (百万) · Daily Output Tokens (M)", 0.3, "B"),
    ]
    inp_row = calc_start + 1
    out_row = calc_start + 2
    yellow = PatternFill("solid", fgColor="FFFF99")

    for i, (label, default, _) in enumerate(labels):
        r = calc_start + 1 + i
        ws2[f"A{r}"].value = label
        ws2[f"A{r}"].font = Font(name="Arial", size=9)
        ws2[f"A{r}"].alignment = Alignment(vertical="center")
        ws2[f"B{r}"].value = default
        ws2[f"B{r}"].fill = yellow
        ws2[f"B{r}"].font = Font(name="Arial", size=9, bold=True, color="0000CC")
        ws2[f"B{r}"].number_format = "0.00"
        ws2[f"B{r}"].alignment = Alignment(horizontal="center")
        ws2.row_dimensions[r].height = 16

    result_row = calc_start + 4
    # Header
    for col, label in [("A","服务商"), ("B","模型"), ("C","日成本 Daily ($)"), ("D","月成本 Monthly ($)"), ("E","年成本 Annual ($)")]:
        c = ws2[f"{col}{result_row}"]
        c.value = label
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="37474F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = header_border()
    ws2.row_dimensions[result_row].height = 18

    # Populate cost rows using formulas referencing inp_row/out_row B cells
    # and the actual prices embedded as values (since they're on another sheet,
    # we just store the price directly and compute cost here)
    selected_models = [
        ("OpenAI",     "GPT-5.4",          2.50,  15.00),
        ("OpenAI",     "GPT-5.2",          1.75,  14.00),
        ("OpenAI",     "GPT-4o mini",       0.15,   0.60),
        ("Anthropic",  "Claude Opus 4.6",   5.00,  25.00),
        ("Anthropic",  "Claude Sonnet 4.6", 3.00,  15.00),
        ("Anthropic",  "Claude Haiku 4.5",  0.80,   4.00),
        ("Google",     "Gemini 3.1 Pro",    2.00,  12.00),
        ("Google",     "Gemini 2.5 Flash",  0.30,   2.50),
        ("Google",     "Gemini 2.5 Flash-Lite", 0.10, 0.40),
        ("DeepSeek",   "DeepSeek V3.2",     0.28,   0.42),
        ("Mistral",    "Mistral Nemo",       0.02,   0.04),
        ("xAI",        "Grok 4.20",         2.00,   6.00),
        ("MoonshotAI", "Kimi K2.5",         0.45,   2.20),
    ]

    cr = result_row + 1
    for provider, model, inp_p, out_p in selected_models:
        bg = PROVIDER_COLORS.get(provider, "FFFFFF")
        rfill = PatternFill("solid", fgColor=bg)
        ws2[f"A{cr}"].value = provider
        ws2[f"B{cr}"].value = model
        # daily = input_M * inp_price + output_M * out_price
        # B{inp_row} = daily input M tokens, B{out_row} = daily output M tokens
        ws2[f"C{cr}"].value = f"=B${inp_row}*{inp_p}+B${out_row}*{out_p}"
        ws2[f"D{cr}"].value = f"=C{cr}*30"
        ws2[f"E{cr}"].value = f"=C{cr}*365"
        for col in "ABCDE":
            c = ws2[f"{col}{cr}"]
            c.fill = rfill
            c.border = thin_border()
            c.font = Font(name="Arial", size=9,
                          color="008000" if col in "CDE" else "1A1A2E")
            c.alignment = Alignment(horizontal="right" if col in "CDE" else "left",
                                    vertical="center")
            if col in "CDE":
                c.number_format = '"$"#,##0.00'
        ws2.row_dimensions[cr].height = 16
        cr += 1

    # Color scale on daily cost column
    ws2.conditional_formatting.add(
        f"C{result_row+1}:C{cr-1}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    )

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    build()
